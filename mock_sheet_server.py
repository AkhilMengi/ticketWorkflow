"""
Standalone mock sheet server — run this separately to receive and inspect
real HTTP POST requests from the agent when SHEET_API_URL is configured.

Run it with:
    python mock_sheet_server.py

It listens on port 9001. All received updates are:
  1. Printed to the console (visible immediately)
  2. Stored in memory and accessible via GET http://localhost:9001/updates
  3. ACTUALLY UPDATE abc.xlsx Excel file
  4. Saved to sheet_updates_log.json on disk for persistence

Set in .env:
  SHEET_API_URL=http://localhost:9001
"""
import json
import uuid
import logging
from datetime import datetime, timezone
from typing import List, Dict, Any
from pathlib import Path

from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse
import uvicorn

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [SHEET-SERVER] %(message)s",
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Mock Sheet Server", version="1.0.0")

# In-memory store of all received updates
_received_updates: List[Dict[str, Any]] = []
LOG_FILE = Path("sheet_updates_log.json")
EXCEL_FILE = Path("abc.xlsx")

# Column mappings
ACCOUNT_ID_COLUMN = 11  # Column K = 11


def _save_to_disk():
    LOG_FILE.write_text(json.dumps(_received_updates, indent=2))


def _update_excel_file(account_id: str, field_name: str, field_value: Any) -> Dict[str, Any]:
    """
    Update the Excel file:
    - Find row where Column K (account_id column) matches
    - Find column with header matching field_name
    - Update that cell
    - Save the file
    """
    if not OPENPYXL_AVAILABLE:
        return {
            "success": False,
            "error": "openpyxl not installed",
            "previous_value": None,
        }
    
    if not EXCEL_FILE.exists():
        return {
            "success": False,
            "error": f"Excel file not found: {EXCEL_FILE}",
            "previous_value": None,
        }
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Read header row (row 1) to find the column index for field_name
        field_column = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value and str(cell.value).lower() == field_name.lower():
                field_column = col_idx
                break
        
        if not field_column:
            return {
                "success": False,
                "error": f"Column '{field_name}' not found in header row",
                "previous_value": None,
            }
        
        # Find the row where Column K matches account_id
        target_row = None
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=ACCOUNT_ID_COLUMN).value
            if cell_value and str(cell_value).strip() == str(account_id).strip():
                target_row = row_idx
                break
        
        if not target_row:
            return {
                "success": False,
                "error": f"Account ID '{account_id}' not found in Column K",
                "previous_value": None,
            }
        
        # Get previous value
        previous_cell = ws.cell(row=target_row, column=field_column)
        previous_value = previous_cell.value
        
        # Update the cell
        previous_cell.value = field_value
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        wb.close()
        
        logger.info("✅ Excel file updated: Row %d, Column %d ('%s')", target_row, field_column, field_name)
        
        return {
            "success": True,
            "previous_value": previous_value,
            "new_value": field_value,
        }
    
    except Exception as exc:
        logger.error("Error updating Excel file: %s", exc)
        return {
            "success": False,
            "error": str(exc),
            "previous_value": None,
        }


# ── POST /api/v1/sheet/update ─────────────────────────────────────────────────

@app.post("/api/v1/sheet/update")
async def update_sheet(request: Request):
    """
    Receive a sheet update request and update the Excel file.
    
    Expected body:
    {
        "sheet_file_name": "abc",
        "account_id": "ACC001",
        "field_name": "meter_config",
        "field_value": "D367_NEW",
        "context": "optional context",
        "updated_by": "intelligent-agent"
    }
    """
    body = await request.json()

    update_id = f"UPDATE-{uuid.uuid4().hex[:8].upper()}"
    updated_at = datetime.now(timezone.utc).isoformat()
    
    # Try to update the Excel file
    excel_result = _update_excel_file(
        account_id=body.get("account_id"),
        field_name=body.get("field_name"),
        field_value=body.get("field_value"),
    )
    
    # Record the update
    update_record = {
        "update_id": update_id,
        "sheet_file_name": body.get("sheet_file_name", "unknown"),
        "account_id": body.get("account_id"),
        "field_name": body.get("field_name"),
        "previous_value": excel_result.get("previous_value"),
        "new_value": excel_result.get("new_value", body.get("field_value")),
        "context": body.get("context", ""),
        "updated_by": body.get("updated_by", "unknown"),
        "updated_at": updated_at,
        "status": "completed" if excel_result.get("success") else "failed",
        "error": excel_result.get("error"),
    }

    _received_updates.append(update_record)
    _save_to_disk()

    # Print clearly so you can see it instantly in the terminal
    logger.info("=" * 70)
    logger.info("✅ SHEET UPDATE RECEIVED")
    logger.info("   update_id       : %s", update_id)
    logger.info("   sheet_file_name : %s", body.get("sheet_file_name"))
    logger.info("   account_id      : %s", body.get("account_id"))
    logger.info("   field_name      : %s", body.get("field_name"))
    logger.info("   previous_value  : %s", excel_result.get("previous_value"))
    logger.info("   new_value       : %s", body.get("field_value"))
    logger.info("   context         : %s", str(body.get("context", ""))[:80])
    logger.info("   updated_by      : %s", body.get("updated_by"))
    logger.info("   excel_status    : %s", "OK" if excel_result.get("success") else f"FAILED - {excel_result.get('error')}")
    logger.info("=" * 70)

    return JSONResponse(
        status_code=200 if excel_result.get("success") else 400,
        content={
            "success": excel_result.get("success", False),
            "update_id": update_id,
            "status": update_record["status"],
            "message": f"Sheet '{body.get('sheet_file_name')}' update: {body.get('field_name')} for account {body.get('account_id')}.",
            "previous_value": excel_result.get("previous_value"),
            "new_value": body.get("field_value"),
            "updated_at": updated_at,
            "error": excel_result.get("error"),
        },
    )


# ── GET /updates ──────────────────────────────────────────────────────────────

@app.get("/updates")
async def list_updates():
    """Return all sheet updates received in this session."""
    return {
        "total": len(_received_updates),
        "updates": _received_updates,
    }


@app.get("/updates/{update_id}")
async def get_update(update_id: str):
    """Look up a specific update by update_id."""
    match = next((u for u in _received_updates if u.get("update_id") == update_id), None)
    if not match:
        return JSONResponse(status_code=404, content={"error": "Update not found"})
    return match


@app.get("/updates/account/{account_id}")
async def get_updates_by_account(account_id: str):
    """Get all updates for a specific account."""
    matches = [u for u in _received_updates if u.get("account_id") == account_id]
    return {
        "account_id": account_id,
        "total": len(matches),
        "updates": matches,
    }


@app.delete("/updates")
async def clear_updates():
    """Clear all stored updates (useful between test runs)."""
    _received_updates.clear()
    if LOG_FILE.exists():
        LOG_FILE.unlink()
    return {"message": "All updates cleared."}


@app.get("/health")
async def health():
    return {"status": "ok", "updates_received": len(_received_updates)}


if __name__ == "__main__":
    logger.info("Mock Sheet Server starting on http://localhost:9001")
    logger.info("Agent should have: SHEET_API_URL=http://localhost:9001")
    uvicorn.run(app, host="0.0.0.0", port=9001, log_level="warning")
